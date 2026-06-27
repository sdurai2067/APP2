# =============================================================
# app.py — FSD2TESTDOC App2
# =============================================================
# PROJECT  : FSD2TESTDOC App2
# PURPOSE  : Convert insurance FSD (PDF/DOCX/TXT) into structured
#            JSON1 and generate quality test cases + Excel output.
#
# FRAMEWORK: Flask (Python micro web server)
#            Runs at http://localhost:5000
#            Browser sends requests → Flask runs functions → returns results
#
# PIPELINE :
#   Step 1 → Upload FSD → /api/step1 → JSON1 (1 Gemini call)
#   Step 2 → JSON1      → /api/step2 → Summary + Test Cases (2 calls)
#   Step 3 → JSON1      → /api/step3 → Gap Analysis (1 call, optional)
#   Export → JSON1+Tests→ /api/export→ Excel (0 calls, pure Python)
#   Chat   → JSON1      → /api/chat  → AI answer (1 call per message)
#
# API CALLS PER SESSION:
#   Core  : 3 calls (step1 + step2 summary + step2 tests)
#   Extra : step3 gap=1, chat=1 per message
#   Excel : 0 calls
#
# LOGGING  : Every action logs to terminal with timestamp + colour
#            Format: [HH:MM:SS] [LABEL] message
#
# HOW TO RUN:
#   1. Create .env with GEMINI_API_KEY=AIzaSy...
#   2. pip install -r requirements.txt
#   3. python app.py
#   4. Open http://localhost:5000
# =============================================================


# ─── STANDARD LIBRARY ─────────────────────────────────────────
import os           # read environment variables and file system
import io           # handle file bytes in memory (no disk write needed)
import re           # regex — strip markdown fences from AI responses
import json         # parse/serialise JSON data
import traceback    # print full error details on exception
from pathlib import Path          # cross-platform file paths
from datetime import datetime     # timestamps for logs and filenames


# ─── THIRD PARTY LIBRARIES ────────────────────────────────────
import requests     # HTTP client — calls Gemini API over internet
import urllib3      # low-level HTTP — used to suppress SSL warnings

# Suppress the InsecureRequestWarning that appears when verify=False
# This is needed because corporate Zscaler proxy intercepts SSL traffic
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Flask web framework
from flask import Flask, request, jsonify, send_file, render_template

# Document parsing libraries
import docx             # python-docx: reads .docx Word files
from pypdf import PdfReader  # pypdf: reads .pdf files page by page

# Excel generation library
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# All AI prompts live in prompts.py — edit that file to tune AI
import prompts


# =============================================================
# LOG HELPER — live coloured terminal output
# =============================================================
# Every step in the app calls log() so you can watch live in terminal
# Colours: STEP=cyan, GEMINI=purple, EXCEL=green, FILE=yellow, ERROR=red
# flush=True forces immediate output even inside long-running calls
# =============================================================
COLOUR = {
    "STEP1":  "\033[96m",   # cyan
    "STEP2":  "\033[96m",   # cyan
    "STEP3":  "\033[96m",   # cyan
    "GEMINI": "\033[95m",   # purple
    "EXCEL":  "\033[92m",   # green
    "FILE":   "\033[93m",   # yellow
    "JSON":   "\033[33m",   # dark yellow
    "HEALTH": "\033[94m",   # blue
    "CHAT":   "\033[94m",   # blue
    "EXPORT": "\033[92m",   # green
    "STARTUP":"\033[96m",   # cyan
    "ERROR":  "\033[91m",   # red
}
RESET = "\033[0m"

def log(label, message):
    """
    Print a timestamped, coloured log line to terminal.

    label   : short identifier shown in brackets e.g. STEP1, GEMINI
    message : what is happening right now
    """
    now   = datetime.now().strftime("%H:%M:%S")  # current time HH:MM:SS
    color = COLOUR.get(label, "\033[0m")          # get colour for label
    # Print: [14:32:01] [STEP1] message here
    print(f"{color}[{now}] [{label}]{RESET} {message}", flush=True)


# =============================================================
# LOAD .env FILE
# =============================================================
# .env is a plain text file holding secret values:
#   GEMINI_API_KEY=AIzaSyXXXXXXXXX
#   GEMINI_MODEL=gemini-2.0-flash
#   PORT=5000
#
# We read it line by line and load into os.environ
# os.environ is a dictionary of system-wide settings
# setdefault() = only set if not already set (won't override)
# =============================================================
def load_env():
    """Read .env file and load KEY=VALUE pairs into os.environ."""
    log("STARTUP", "Reading .env file...")
    p = Path(".env")

    # Warn if .env missing — app cannot call Gemini without key
    if not p.exists():
        log("ERROR", "No .env file found!")
        log("ERROR", "Create .env with: GEMINI_API_KEY=AIzaSy...")
        return

    # Read all lines, skip blanks and comments (#)
    for line in p.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        # Split "KEY=VALUE" at first = sign
        k, _, v = line.partition("=")
        os.environ.setdefault(k.strip(), v.strip())
        # Show key loaded (mask value for security — show only first 6 chars)
        log("STARTUP", f"  Loaded: {k.strip()} = {v[:6]}...")

    log("STARTUP", ".env loaded OK")


# Run load_env when module starts (before Flask runs)
load_env()


# =============================================================
# CONFIGURATION
# =============================================================
# Read values from os.environ into Python constants
# .get("KEY", "default") = use "default" if KEY not found
# =============================================================
API_KEY = os.environ.get("GEMINI_API_KEY", "")        # Gemini API key
MODEL   = os.environ.get("GEMINI_MODEL",   "gemini-2.0-flash")  # which model
PORT    = int(os.environ.get("PORT", 5000))            # web server port

# Gemini REST endpoint — the URL we POST to for every AI request
# No SDK needed — plain HTTP POST with JSON payload
GEMINI_URL = (
    f"https://generativelanguage.googleapis.com/v1beta/models/"
    f"{MODEL}:generateContent"
)

# Results folder — JSON1 files and Excel files saved here
Path("results").mkdir(exist_ok=True)

# Log configuration summary at startup
log("STARTUP", f"Project  : FSD2TESTDOC App2")
log("STARTUP", f"Model    : {MODEL}")
log("STARTUP", f"Port     : {PORT}")
log("STARTUP", f"API Key  : {'SET ✓' if API_KEY else 'MISSING ✗'}")
log("STARTUP", f"Endpoint : {GEMINI_URL[:55]}...")
log("STARTUP", f"Results  : ./results/")


# =============================================================
# FLASK APP
# =============================================================
# Flask(__name__) creates the web server
# template_folder="." tells Flask index.html is in same folder
# MAX_CONTENT_LENGTH limits upload file size to 50MB
# =============================================================
app = Flask(__name__, template_folder=".")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB upload limit
log("STARTUP", "Flask app initialised")


# =============================================================
# GEMINI API CALLER
# =============================================================
# Single function used for ALL Gemini API calls in this app.
# All 4 AI calls (json1, summary, tests, gap) go through here.
#
# REQUEST FORMAT (what Gemini expects):
#   POST https://.../{model}:generateContent?key={api_key}
#   Body: { "contents": [{ "parts": [{ "text": "prompt" }] }],
#            "generationConfig": { temperature, maxOutputTokens } }
#
# RESPONSE FORMAT (what Gemini returns):
#   { "candidates": [{ "content": { "parts": [{ "text": "..." }] } }] }
#
# SETTINGS:
#   temperature=0.2  : very low = consistent factual output
#   maxOutputTokens=8192 : enough for 15 detailed test cases
#   verify=False     : bypass Zscaler corporate SSL proxy
#   timeout=180      : wait up to 3 minutes for large responses
# =============================================================
def call_gemini(prompt, label="GEMINI"):
    """
    Send a prompt to Gemini API and return the text response.

    prompt : fully formatted prompt string to send
    label  : short name shown in logs (e.g. STEP1, STEP2-TESTS)

    Returns : string — Gemini's text response
    Raises  : RuntimeError on quota, invalid key, or connection error
    """

    # Cannot call Gemini without an API key
    if not API_KEY:
        log("ERROR", "GEMINI_API_KEY is not set in .env!")
        raise RuntimeError(
            "GEMINI_API_KEY not set. "
            "Add to .env: GEMINI_API_KEY=AIzaSy..."
        )

    log(label, f"Sending {len(prompt):,} chars to Gemini...")

    # Build the JSON payload Gemini expects
    payload = {
        "contents": [
            {"parts": [{"text": prompt}]}  # the prompt goes here
        ],
        "generationConfig": {
            "temperature":     0.2,    # 0=robotic, 1=creative, 0.2=factual
            "maxOutputTokens": 8192,   # max length of response
            "topP":            0.9,    # controls vocabulary diversity
        }
    }

    # Make the HTTP POST request to Gemini
    # API key goes in URL query param (?key=...) — Gemini's requirement
    # verify=False bypasses SSL certificate check (needed for Zscaler)
    try:
        resp = requests.post(
            f"{GEMINI_URL}?key={API_KEY}",
            headers={"Content-Type": "application/json"},
            json=payload,
            timeout=180,    # 3 minute timeout for large responses
            verify=False,   # bypass corporate SSL proxy (Zscaler)
        )
    except requests.exceptions.Timeout:
        log("ERROR", "Gemini request timed out after 3 minutes")
        raise RuntimeError("Request timed out. Try again or use shorter FSD.")

    except requests.exceptions.ConnectionError:
        log("ERROR", "Cannot connect to Gemini — check internet")
        raise RuntimeError(
            "Cannot reach Gemini. Check internet connection. "
            "If on corporate network, SSL bypass is already set."
        )

    # Log the HTTP status code
    log(label, f"HTTP {resp.status_code} received")

    # Handle specific error status codes with clear messages
    if resp.status_code == 429:
        # Too Many Requests — free quota exhausted
        log("ERROR", "QUOTA EXCEEDED (429) — free limit hit!")
        raise RuntimeError(
            "Quota exceeded (429). "
            "Wait 60 seconds OR create new key at "
            "https://aistudio.google.com/apikey"
        )

    if resp.status_code == 403:
        # Forbidden — API key invalid or expired
        log("ERROR", "API KEY INVALID (403)")
        raise RuntimeError(
            "API key rejected (403). "
            "Get fresh key at https://aistudio.google.com/apikey"
        )

    if resp.status_code == 400:
        # Bad Request — usually wrong model name in .env
        log("ERROR", f"BAD REQUEST (400) — model name wrong? Current: {MODEL}")
        raise RuntimeError(
            f"Bad request (400). Check GEMINI_MODEL in .env. "
            f"Current: {MODEL}. Try: gemini-2.0-flash"
        )

    if resp.status_code != 200:
        # Any other unexpected error
        log("ERROR", f"Unexpected error: HTTP {resp.status_code}")
        raise RuntimeError(
            f"Gemini error {resp.status_code}: {resp.text[:200]}"
        )

    # Parse the nested JSON response structure
    # Path: response["candidates"][0]["content"]["parts"][0]["text"]
    try:
        data = resp.json()
        text = data["candidates"][0]["content"]["parts"][0]["text"]
        log(label, f"Response OK: {len(text):,} characters received")
        return text

    except (KeyError, IndexError) as e:
        log("ERROR", f"Unexpected response format: {e}")
        raise RuntimeError(f"Could not parse Gemini response: {data}")


# =============================================================
# FILE TEXT EXTRACTOR
# =============================================================
# Reads the uploaded file and returns plain text.
# Supports PDF, DOCX, TXT.
#
# PDF  : pypdf reads page by page
#        SKIPS first 2 pages (cover page + author/version page)
#        Cover pages are administrative — functional content
#        starts from page 3 onwards in most insurance FSDs
#
# DOCX : python-docx reads each paragraph
#        SKIPS first 3 paragraphs (document title, author, date)
#
# TXT  : simple UTF-8 decode
# =============================================================
def extract_text(file_storage):
    """
    Extract plain text from an uploaded PDF, DOCX, or TXT file.

    file_storage : Flask FileStorage object from request.files["fsd"]

    Returns : (text_string, page_count) tuple
    """
    name = file_storage.filename.lower()  # lowercase for extension check
    raw  = file_storage.read()            # read entire file as raw bytes
    log("FILE", f"Reading: {file_storage.filename} ({len(raw):,} bytes)")

    # ── PDF ──────────────────────────────────────────────────
    if name.endswith(".pdf"):
        reader      = PdfReader(io.BytesIO(raw))  # open PDF from bytes
        total_pages = len(reader.pages)
        log("FILE", f"PDF: {total_pages} pages found")

        pages = []
        for i, page in enumerate(reader.pages):

            # Skip pages 0 and 1 (first 2 pages = cover + author)
            if i < 2:
                log("FILE", f"  Skipping page {i+1} (cover/author)")
                continue

            # Extract text from this page
            text = page.extract_text()
            if text and text.strip():
                pages.append(f"[PAGE {i+1}]\n{text}")
                log("FILE", f"  Page {i+1}: {len(text):,} chars")
            else:
                log("FILE", f"  Page {i+1}: empty or image-based (skipped)")

        result = "\n\n".join(pages)
        log("FILE", f"PDF extraction done: {len(result):,} total chars")
        return result, total_pages

    # ── DOCX ─────────────────────────────────────────────────
    elif name.endswith(".docx"):
        doc   = docx.Document(io.BytesIO(raw))   # open Word doc from bytes
        paras = [p.text for p in doc.paragraphs  # get all non-empty paragraphs
                 if p.text.strip()]
        log("FILE", f"DOCX: {len(paras)} paragraphs found")

        # Skip first 3 paragraphs (title, author name, date line)
        content = paras[3:] if len(paras) > 5 else paras
        log("FILE", f"  Skipping first 3 paragraphs (title/author/date)")

        result = "\n".join(content)
        log("FILE", f"DOCX extraction done: {len(result):,} total chars")
        return result, len(paras)

    # ── TXT ──────────────────────────────────────────────────
    elif name.endswith(".txt"):
        result = raw.decode("utf-8", errors="ignore")
        log("FILE", f"TXT decoded: {len(result):,} chars")
        return result, 1

    # ── FALLBACK ─────────────────────────────────────────────
    else:
        log("FILE", f"Unknown type — attempting UTF-8 decode")
        result = raw.decode("utf-8", errors="ignore")
        log("FILE", f"Fallback decode: {len(result):,} chars")
        return result, 1


# =============================================================
# JSON PARSER
# =============================================================
# AI models sometimes wrap JSON in markdown code fences:
#   ```json
#   { "key": "value" }
#   ```
#
# This function:
# 1. Strips the markdown fences
# 2. Finds the first [ or { (start of JSON structure)
# 3. Walks forward tracking nested brackets
# 4. Extracts the complete JSON substring
# 5. Parses it with json.loads()
# =============================================================
def parse_json(text):
    """
    Extract and parse valid JSON from an AI text response.

    text : raw string from Gemini (may contain markdown fences)

    Returns : parsed dict or list, or None if no valid JSON found
    """
    log("JSON", f"Parsing AI response ({len(text):,} chars)")

    # Step 1: Remove markdown code fences ```json ... ```
    text = re.sub(r"```(?:json)?", "", text).strip().rstrip("`").strip()

    # Step 2: Find first [ or { character — start of JSON
    start = next((i for i, c in enumerate(text) if c in "[{"), -1)
    if start == -1:
        log("JSON", "WARNING: No JSON structure found in response")
        return None

    # Step 3: Walk forward tracking bracket depth
    depth, in_str, esc = 0, False, False
    opener = text[start]
    closer = "]" if opener == "[" else "}"

    for i in range(start, len(text)):
        ch = text[i]
        if esc:        esc = False;     continue   # escaped char
        if ch == "\\":  esc = True;     continue   # escape sequence start
        if ch == '"':   in_str = not in_str        # toggle string mode
        if not in_str:                             # only count brackets outside strings
            if ch == opener: depth += 1
            elif ch == closer:
                depth -= 1
                if depth == 0:
                    # Found complete JSON structure — parse it
                    try:
                        result = json.loads(text[start:i+1])
                        log("JSON", "JSON parsed successfully")
                        return result
                    except json.JSONDecodeError as e:
                        log("JSON", f"Parse error at position {i}: {e}")
                        break

    # Last resort: try parsing the entire cleaned text
    try:
        result = json.loads(text)
        log("JSON", "JSON parsed (fallback method)")
        return result
    except Exception:
        log("JSON", "ERROR: Could not extract valid JSON from response")
        return None


# =============================================================
# EXCEL BUILDER — 3 sheets, ZERO API calls
# =============================================================
# Builds a professionally formatted .xlsx workbook from pipeline data.
# Sheet 1: Test Cases  — colour-coded by scenario type
# Sheet 2: FSD Summary — metadata, flow, rules, risks
# Sheet 3: Coverage    — scenario/priority breakdown stats
#
# Uses openpyxl library entirely. No Gemini call needed.
# =============================================================
def build_excel(tcs, summary, json1, fname):
    """
    Build a 3-sheet Excel workbook from test cases and summary.

    tcs     : list of test case dicts from Gemini
    summary : summary dict from Gemini
    json1   : JSON1 dict (for metadata in Sheet 2)
    fname   : original FSD filename (used in title banner)

    Returns : bytes of the complete .xlsx file
    """
    log("EXCEL", f"Building Excel: {len(tcs)} test cases, file={fname}")

    wb = Workbook()  # create new empty workbook

    # ── Colour constants ──────────────────────────────────────
    NAVY  = "0F172A"   # dark navy for titles and headers
    BLUE  = "2563EB"   # bright blue for sub-banners
    GRAY  = "374151"   # dark gray for column headers
    WHITE = "FFFFFF"   # white for cell backgrounds

    # Background and text colours per scenario type
    SCN_BG = {
        "POSITIVE":   "D1FAE5",   # light green
        "NEGATIVE":   "FEE2E2",   # light red
        "BOUNDARY":   "FEF3C7",   # light yellow
        "EXCEPTION":  "FCE7F3",   # light pink
        "REGRESSION": "EDE9FE",   # light purple
    }
    SCN_FG = {
        "POSITIVE":   "065F46",   # dark green text
        "NEGATIVE":   "991B1B",   # dark red text
        "BOUNDARY":   "92400E",   # dark amber text
        "EXCEPTION":  "9D174D",   # dark pink text
        "REGRESSION": "4C1D95",   # dark purple text
    }

    # ── Border styles ─────────────────────────────────────────
    thin = Border(
        left=Side(style="thin",   color="CBD5E1"),
        right=Side(style="thin",  color="CBD5E1"),
        top=Side(style="thin",    color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    thick = Border(
        left=Side(style="medium",   color=NAVY),
        right=Side(style="medium",  color=NAVY),
        top=Side(style="medium",    color=NAVY),
        bottom=Side(style="medium", color=NAVY),
    )

    def hc(ws, r, c, v, bg=NAVY, fg=WHITE):
        """Apply header cell: bold, coloured background, centred."""
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(name="Arial", size=10, bold=True, color=fg)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thick
        return cell

    def dc(ws, r, c, v, bg=WHITE, sz=9, ha="left"):
        """Apply data cell: normal font, light background."""
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(name="Arial", size=sz, color="1F2937")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=ha, vertical="top", wrap_text=True)
        cell.border    = thin
        return cell

    # ─────────────────────────────────────────────────────────
    # SHEET 1 — TEST CASES
    # ─────────────────────────────────────────────────────────
    log("EXCEL", "Writing Sheet 1: Test Cases")
    ws = wb.active
    ws.title = "Test Cases"
    ws.sheet_view.showGridLines = False   # cleaner look without grid

    # Title banner — row 1 merged across all columns
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 38
    t = ws["A1"]
    t.value     = f"TEST CASE DOCUMENT  |  FSD2TESTDOC App2  |  {fname}"
    t.font      = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill      = PatternFill("solid", fgColor=NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")

    # Sub-banner — row 2 with generation metadata
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 18
    s2 = ws["A2"]
    s2.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        f"   |   Total: {len(tcs)} test cases"
        f"   |   Model: {MODEL}"
    )
    s2.font      = Font(name="Arial", size=9, color=WHITE)
    s2.fill      = PatternFill("solid", fgColor=BLUE)
    s2.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers — row 3
    hdrs   = ["S.No", "TC No", "FS Ref", "Test Description", "Scenario",
              "Steps", "Expected Result", "Pass/Fail", "Priority"]
    widths = [6, 10, 7, 30, 13, 52, 48, 11, 10]
    ws.row_dimensions[3].height = 26
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        hc(ws, 3, ci, h, bg=GRAY)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows — one per test case
    for ri, tc in enumerate(tcs):
        r   = ri + 4
        ws.row_dimensions[r].height = 78        # tall row for multi-line steps
        alt = "F8FAFF" if ri % 2 else WHITE     # alternating row colour
        scn = str(tc.get("scenario", "POSITIVE")).upper()

        vals = [
            tc.get("sno",              ri + 1),
            tc.get("tc_no",            f"TC-{ri+1:03d}"),
            tc.get("fs_ref",           "FS1"),
            tc.get("test_description", ""),
            scn,
            tc.get("steps",            ""),
            tc.get("expected_result",  ""),
            tc.get("pass_fail",        ""),
            tc.get("priority",         "HIGH"),
        ]

        for ci, v in enumerate(vals, 1):
            if ci == 5:
                # Scenario column — colour coded by type
                cell = ws.cell(row=r, column=ci, value=v)
                cell.font      = Font(name="Arial", size=9, bold=True,
                                      color=SCN_FG.get(scn, "1F2937"))
                cell.fill      = PatternFill("solid",
                                             fgColor=SCN_BG.get(scn, alt))
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center", wrap_text=True)
                cell.border    = thin
            else:
                dc(ws, r, ci, v, bg=alt,
                   ha="center" if ci in [1, 8, 9] else "left")

    ws.freeze_panes = "A4"  # freeze title + header rows
    log("EXCEL", f"  Sheet 1 done: {len(tcs)} rows written")

    # ─────────────────────────────────────────────────────────
    # SHEET 2 — FSD SUMMARY
    # ─────────────────────────────────────────────────────────
    log("EXCEL", "Writing Sheet 2: FSD Summary")
    ws2 = wb.create_sheet("FSD Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 80

    # Summary title
    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 34
    b = ws2["A1"]
    b.value     = "FSD SUMMARY — FSD2TESTDOC App2 — INGENIUM / POLICY ADMIN"
    b.font      = Font(name="Arial", size=13, bold=True, color=WHITE)
    b.fill      = PatternFill("solid", fgColor=NAVY)
    b.alignment = Alignment(horizontal="center", vertical="center")

    def write_section(ws, row, title, items, bg="DBEAFE"):
        """Write a coloured section header followed by a bullet list of items."""
        items = [str(i) for i in (items or [])]
        ws.merge_cells(f"A{row}:B{row}")
        ws.row_dimensions[row].height = 22
        sc        = ws.cell(row=row, column=1, value=f"  {title}")
        sc.font   = Font(name="Arial", size=10, bold=True, color=NAVY)
        sc.fill   = PatternFill("solid", fgColor=bg)
        sc.border = Border(bottom=Side(style="medium", color=NAVY))

        for i, item in enumerate(items):
            r2  = row + 1 + i
            ws.row_dimensions[r2].height = 18
            bg2 = WHITE if i % 2 == 0 else "F9FAFB"
            # Bullet in column A
            lc        = ws.cell(row=r2, column=1, value="  •")
            lc.font   = Font(name="Arial", size=9)
            lc.fill   = PatternFill("solid", fgColor=bg2)
            lc.border = thin
            # Item text in column B
            vc            = ws.cell(row=r2, column=2, value=f"  {item}")
            vc.font       = Font(name="Arial", size=9)
            vc.alignment  = Alignment(wrap_text=True, vertical="top")
            vc.fill       = PatternFill("solid", fgColor=bg2)
            vc.border     = thin

        return row + 1 + len(items) + 1  # next available row

    # Write summary sections in order with different colours
    s   = summary or {}
    row = 2
    row = write_section(ws2, row, "CENTER POINT",
                        [s.get("center_point", "")],            bg="DBEAFE")
    if s.get("transaction_types"):
        row = write_section(ws2, row, "TRANSACTION TYPES",
                            s["transaction_types"],              bg="EDE9FE")
    if s.get("business_flow"):
        row = write_section(ws2, row, "BUSINESS PROCESS FLOW",
                            s["business_flow"],                  bg="D1FAE5")
    if s.get("key_rules"):
        row = write_section(ws2, row, "KEY BUSINESS RULES",
                            s["key_rules"],                      bg="FEF3C7")
    if s.get("critical_inputs"):
        row = write_section(ws2, row, "CRITICAL INPUT FIELDS",
                            s["critical_inputs"],                bg="CCFBF1")
    if s.get("critical_outputs"):
        row = write_section(ws2, row, "CRITICAL OUTPUT FIELDS",
                            s["critical_outputs"],               bg="FEF9C3")
    if s.get("integration_map"):
        row = write_section(ws2, row, "INTEGRATION TOUCHPOINTS",
                            s["integration_map"],                bg="FCE7F3")
    if s.get("test_data_needed"):
        row = write_section(ws2, row, "TEST DATA REQUIRED (INGENIUM)",
                            s["test_data_needed"],               bg="FFF7ED")
    if s.get("risks"):
        row = write_section(ws2, row, "RISKS & ASSUMPTIONS",
                            s.get("risks", []) + s.get("assumptions", []),
                                                                 bg="FEE2E2")
    ws2.freeze_panes = "A2"
    log("EXCEL", "  Sheet 2 done")

    # ─────────────────────────────────────────────────────────
    # SHEET 3 — COVERAGE MATRIX
    # ─────────────────────────────────────────────────────────
    log("EXCEL", "Writing Sheet 3: Coverage Matrix")
    ws3 = wb.create_sheet("Coverage Matrix")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    ws3.row_dimensions[1].height = 28
    cm = ws3["A1"]
    cm.value     = "TEST COVERAGE MATRIX — FSD2TESTDOC App2"
    cm.font      = Font(name="Arial", size=12, bold=True, color=WHITE)
    cm.fill      = PatternFill("solid", fgColor=NAVY)
    cm.alignment = Alignment(horizontal="center", vertical="center")

    # Coverage table headers
    ch = ["Scenario", "Count", "% Coverage", "TC IDs", "Risk Level", "Status"]
    cw = [16, 8, 12, 52, 12, 10]
    for ci, (cn, cw2) in enumerate(zip(ch, cw), 1):
        hc(ws3, 2, ci, cn, bg=GRAY)
        ws3.column_dimensions[get_column_letter(ci)].width = cw2

    # Count test cases per scenario type
    counts, tc_ids = {}, {}
    for tc in tcs:
        scn = str(tc.get("scenario", "POSITIVE")).upper()
        counts[scn] = counts.get(scn, 0) + 1
        tc_ids.setdefault(scn, []).append(str(tc.get("tc_no", "")))

    total = len(tcs) or 1

    # Risk level associated with each scenario type
    risk_level = {
        "POSITIVE":   "LOW",
        "NEGATIVE":   "HIGH",
        "BOUNDARY":   "MEDIUM",
        "EXCEPTION":  "HIGH",
        "REGRESSION": "MEDIUM",
    }

    # One row per scenario type
    for ri, (scn, cnt) in enumerate(counts.items(), 3):
        ws3.row_dimensions[ri].height = 18
        pct  = f"{round(cnt / total * 100, 1)}%"
        ids  = ", ".join(tc_ids.get(scn, []))
        row_data = [scn, cnt, pct, ids, risk_level.get(scn, "MED"), "PENDING"]

        for ci, v in enumerate(row_data, 1):
            c2 = ws3.cell(row=ri, column=ci, value=v)
            c2.font = Font(name="Arial", size=9,
                           bold=(ci == 1),
                           color=SCN_FG.get(scn, "1F2937") if ci == 1 else "1F2937")
            c2.fill = PatternFill("solid", fgColor=SCN_BG.get(scn, WHITE))
            c2.alignment = Alignment(
                horizontal="center" if ci != 4 else "left",
                vertical="center", wrap_text=True)
            c2.border = thin

    # Total summary row at bottom
    tr = len(counts) + 3
    for ci, v in enumerate(["TOTAL", total, "100%", "", "", ""], 1):
        c3 = ws3.cell(row=tr, column=ci, value=v)
        c3.font      = Font(name="Arial", size=10, bold=True, color=WHITE)
        c3.fill      = PatternFill("solid", fgColor=BLUE)
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c3.border    = thick

    ws3.freeze_panes = "A3"
    log("EXCEL", "  Sheet 3 done")

    # Save entire workbook to bytes and return
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log("EXCEL", f"Workbook complete: {buf.getbuffer().nbytes:,} bytes")
    return buf.read()


# =============================================================
# FLASK ROUTES
# =============================================================
# Each @app.route defines a URL endpoint.
# When browser sends request to that URL → Flask runs the function.
# GET  = browser requesting data (no body)
# POST = browser sending data (file upload or JSON body)
# =============================================================

@app.route("/")
def index():
    """
    Serve the frontend page.
    When user opens http://localhost:5000, this sends index.html.
    render_template() looks for index.html in template_folder (".")
    """
    log("HEALTH", "GET / — serving index.html")
    return render_template("index.html")


@app.route("/api/health")
def health():
    """
    Health check — verify Gemini is reachable.
    Called automatically when page loads.
    Browser uses response to show green or red status dot.

    Returns JSON: { ok: true/false, model: "gemini-..." }
    """
    log("HEALTH", "Health check requested")

    if not API_KEY:
        log("HEALTH", "FAIL — API key not set")
        return jsonify({"ok": False, "reason": "GEMINI_API_KEY not set in .env"})

    try:
        # Minimal 2-token ping just to verify connectivity
        call_gemini("Say OK", label="HEALTH")
        log("HEALTH", "Gemini reachable — status GREEN ✓")
        return jsonify({"ok": True, "model": MODEL})

    except Exception as e:
        log("HEALTH", f"Gemini unreachable: {e}")
        return jsonify({"ok": False, "reason": str(e)})


@app.route("/api/step1", methods=["POST"])
def step1():
    """
    STEP 1 — Upload FSD → Extract JSON1

    Accepts : multipart form with field "fsd" containing the file
    Process :
      1. Read file bytes from upload
      2. Extract text (skip cover pages)
      3. Send text to Gemini with FSD_TO_JSON1 prompt
      4. Parse JSON1 from response
      5. Save JSON1 to results/ folder
      6. Return JSON1 to browser for storage

    JSON1 is the single source of truth for all downstream calls.
    The original FSD text is NOT sent to Gemini again after this step.

    Returns JSON:
      ok       : true/false
      fname    : original filename
      preview  : first 800 chars of extracted text
      json1    : complete JSON1 structure
      sections : number of FS sections found
    """
    log("STEP1", "=" * 50)
    log("STEP1", "STEP 1 — FSD Upload → JSON1 Extraction")
    log("STEP1", "=" * 50)

    # Check file was included in the request
    if "fsd" not in request.files:
        log("STEP1", "ERROR: No file in request")
        return jsonify({"error": "No file uploaded"}), 400

    f     = request.files["fsd"]
    fname = f.filename
    log("STEP1", f"File received: {fname}")

    # Extract text from the uploaded file
    text, pages = extract_text(f)

    if not text.strip():
        log("STEP1", "ERROR: No text could be extracted from file")
        return jsonify({
            "error": "No text extracted. Save FSD as DOCX or TXT and try again."
        }), 400

    log("STEP1", f"Text extracted: {len(text):,} chars from {pages} pages")
    log("STEP1", f"Sending first {min(9000, len(text)):,} chars to Gemini...")

    try:
        # ── GEMINI CALL 1 ────────────────────────────────────
        # FSD text → JSON1
        # This is the only time the raw FSD text is sent to AI
        raw = call_gemini(
            prompts.build(prompts.PROMPT_JSON1, fsd_text=text[:9000]),
            label="STEP1"
        )

        # Parse the JSON1 from Gemini's text response
        log("STEP1", "Parsing JSON1 structure...")
        json1 = parse_json(raw)

        if not json1:
            log("STEP1", "ERROR: JSON1 parse failed")
            return jsonify({
                "error": "AI could not structure the FSD. "
                         "Try saving as DOCX format for better text extraction."
            }), 500

        # Count and log the FS sections found
        sections = json1.get("sections", {})
        log("STEP1", f"JSON1 SUCCESS: {len(sections)} FS sections")
        for k, v in sections.items():
            title = v.get("section_title", "")
            rules = len(v.get("business_rules", []))
            calcs = len(v.get("calculations", []))
            errs  = len(v.get("error_conditions", []))
            log("STEP1", f"  {k}: '{title}' | {rules} rules | {calcs} calcs | {errs} errors")

        # Save JSON1 to results folder for reference and debugging
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        jpath = f"results/JSON1_{Path(fname).stem}_{ts}.json"
        Path(jpath).write_text(json.dumps(json1, indent=2), encoding="utf-8")
        log("STEP1", f"JSON1 saved to: {jpath}")
        log("STEP1", "STEP 1 COMPLETE ✓")

        return jsonify({
            "ok":       True,
            "fname":    fname,
            "preview":  text[:800],
            "json1":    json1,
            "sections": len(sections),
        })

    except Exception as e:
        traceback.print_exc()
        log("STEP1", f"ERROR: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/step2", methods=["POST"])
def step2():
    """
    STEP 2 — JSON1 → Summary + Test Cases

    Accepts : JSON body with field "json1"
    Process :
      1. Receive JSON1 from browser (no file re-read)
      2. CALL 2A: JSON1 → Summary (business flow, rules, etc.)
      3. CALL 2B: JSON1 → Test Cases (12-15 Ingenium-aware cases)
      4. Return both to browser

    Two separate Gemini calls so each gets full token budget.
    Combining them would reduce quality of both outputs.

    Returns JSON:
      ok       : true/false
      summary  : structured summary dict
      tcs      : list of test case dicts (12-15)
      tc_count : total test cases generated
    """
    log("STEP2", "=" * 50)
    log("STEP2", "STEP 2 — JSON1 → Summary + Test Cases")
    log("STEP2", "=" * 50)

    data  = request.get_json() or {}
    json1 = data.get("json1")

    if not json1:
        log("STEP2", "ERROR: No JSON1 received. Run Step 1 first.")
        return jsonify({"error": "No JSON1. Run Step 1 first."}), 400

    # Convert JSON1 dict to string for prompt (limit to 8000 chars)
    j1str = json.dumps(json1, indent=2)[:8000]
    log("STEP2", f"JSON1 received: {len(j1str):,} chars")

    try:
        # ── GEMINI CALL 2A — SUMMARY ─────────────────────────
        log("STEP2", "CALL 2A: Generating Summary...")
        sum_raw = call_gemini(
            prompts.build(prompts.PROMPT_SUMMARY, json1=j1str),
            label="STEP2-SUMMARY"
        )
        summary = parse_json(sum_raw) or {}
        log("STEP2", f"Summary OK: {len(summary)} fields extracted")

        # ── GEMINI CALL 2B — TEST CASES ──────────────────────
        log("STEP2", "CALL 2B: Generating Test Cases...")
        tc_raw = call_gemini(
            prompts.build(prompts.PROMPT_TESTS, json1=j1str),
            label="STEP2-TESTS"
        )
        tcs = parse_json(tc_raw) or []

        # Ensure sequential sno numbers
        for i, tc in enumerate(tcs):
            tc["sno"] = i + 1

        # Log each test case generated
        log("STEP2", f"Tests OK: {len(tcs)} test cases")
        for tc in tcs:
            log("STEP2", (
                f"  {tc.get('tc_no','?')} "
                f"[{tc.get('scenario','?')}] "
                f"[{tc.get('fs_ref','?')}] "
                f"{tc.get('test_description','')[:55]}"
            ))

        log("STEP2", "STEP 2 COMPLETE ✓")

        return jsonify({
            "ok":       True,
            "summary":  summary,
            "tcs":      tcs,
            "tc_count": len(tcs),
        })

    except Exception as e:
        traceback.print_exc()
        log("STEP2", f"ERROR: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/step3", methods=["POST"])
def step3():
    """
    STEP 3 — JSON1 → Gap Analysis (optional)

    Accepts : JSON body with field "json1"
    Process :
      1. Receive JSON1 from browser
      2. CALL 3: JSON1 → Gap Analysis
      3. Return quality score and gap register

    Optional — only runs when user clicks Gap Analysis button.
    Not auto-triggered to save quota.

    Returns JSON:
      ok            : true/false
      quality_score : integer 0-100
      score_reason  : why this score
      gaps          : list of gap objects (gap_id, fs_ref, missing, impact, question)
    """
    log("STEP3", "=" * 50)
    log("STEP3", "STEP 3 — JSON1 → Gap Analysis")
    log("STEP3", "=" * 50)

    data  = request.get_json() or {}
    json1 = data.get("json1")

    if not json1:
        log("STEP3", "ERROR: No JSON1. Run Step 1 first.")
        return jsonify({"error": "No JSON1. Run Step 1 first."}), 400

    try:
        # ── GEMINI CALL 3 — GAP ANALYSIS ─────────────────────
        log("STEP3", "CALL 3: Running Gap Analysis...")
        raw = call_gemini(
            prompts.build(
                prompts.PROMPT_GAP,
                json1=json.dumps(json1, indent=2)[:7000]
            ),
            label="STEP3"
        )
        gap = parse_json(raw) or {"quality_score": 0, "gaps": []}

        log("STEP3", f"Quality score: {gap.get('quality_score', 0)}/100")
        log("STEP3", f"Gaps found: {len(gap.get('gaps', []))}")
        for g in gap.get("gaps", []):
            log("STEP3", f"  {g.get('gap_id','')} [{g.get('fs_ref','')}] {g.get('missing','')[:60]}")

        log("STEP3", "STEP 3 COMPLETE ✓")
        return jsonify({"ok": True, **gap})

    except Exception as e:
        traceback.print_exc()
        log("STEP3", f"ERROR: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/export", methods=["POST"])
def export():
    """
    EXPORT — Build Excel and stream to browser (ZERO API calls)

    Accepts : JSON body with tcs, summary, json1, fname
    Process :
      1. Receive pipeline data from browser
      2. Build 3-sheet Excel workbook using openpyxl (pure Python)
      3. Save copy to results/ folder
      4. Stream .xlsx file to browser as download

    No Gemini call — 100% Python. Uses ZERO quota.

    Returns : .xlsx file download
    """
    log("EXPORT", "=" * 50)
    log("EXPORT", "EXPORT — Building Excel (0 API calls)")
    log("EXPORT", "=" * 50)

    data  = request.get_json() or {}
    tcs   = data.get("tcs",     [])
    sum_  = data.get("summary", {})
    json1 = data.get("json1",   {})
    fname = data.get("fname",   "FSD")

    log("EXPORT", f"Test cases: {len(tcs)}, File: {fname}")

    if not tcs:
        log("EXPORT", "ERROR: No test cases to export")
        return jsonify({"error": "No test cases. Run Step 2 first."}), 400

    try:
        # Build the Excel workbook
        xlsx = build_excel(tcs, sum_, json1, fname)

        # Generate timestamped output filename
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        oname = f"TestDoc_{Path(fname).stem}_{ts}.xlsx"

        # Save a copy to results folder
        Path(f"results/{oname}").write_bytes(xlsx)
        log("EXPORT", f"Saved to: results/{oname}")
        log("EXPORT", "EXPORT COMPLETE ✓ — streaming to browser")

        # Stream the file to browser as a download
        return send_file(
            io.BytesIO(xlsx),
            mimetype=(
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=oname,
        )

    except Exception as e:
        traceback.print_exc()
        log("EXPORT", f"ERROR: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/chat", methods=["POST"])
def chat():
    """
    CHAT — Answer questions using JSON1 as context (optional)

    Accepts : JSON body with json1 and request fields
    Process :
      1. Receive JSON1 + user message from browser
      2. Send both to Gemini with CHAT prompt
      3. Return AI response

    If AI returns JSON with test_cases array, browser detects it
    and offers user the option to apply them to the Test Cases tab.

    Returns JSON:
      ok       : true/false
      response : AI text response
      parsed   : parsed JSON if AI returned structured data
    """
    log("CHAT", "Chat message received")

    data     = request.get_json() or {}
    json1    = data.get("json1", {})
    user_req = data.get("request", "")

    log("CHAT", f"User: {user_req[:80]}")

    if not user_req.strip():
        return jsonify({"error": "No message provided"}), 400

    try:
        # ── GEMINI CALL — CHAT ────────────────────────────────
        raw = call_gemini(
            prompts.build(
                prompts.PROMPT_CHAT,
                json1=json.dumps(json1, indent=2)[:5000],
                user_request=user_req,
            ),
            label="CHAT"
        )
        parsed = parse_json(raw)   # check if response contains JSON
        log("CHAT", f"Response: {len(raw):,} chars")
        return jsonify({"ok": True, "response": raw, "parsed": parsed})

    except Exception as e:
        traceback.print_exc()
        log("CHAT", f"ERROR: {e}")
        return jsonify({"error": str(e)}), 500


# =============================================================
# ENTRY POINT
# =============================================================
# This block runs when you type: python app.py
# It prints a startup summary then starts the Flask server.
# debug=False = stable production-like mode
# host=0.0.0.0 = accessible from any device on your network
# =============================================================
if __name__ == "__main__":
    print()
    print("=" * 58)
    print("  FSD2TESTDOC App2 — FSD Intelligence Pipeline")
    print(f"  Open browser: http://localhost:{PORT}")
    print(f"  Model        : {MODEL}")
    print(f"  API Key      : {'SET ✓' if API_KEY else 'MISSING — add to .env!'}")
    print(f"  Results      : ./results/")
    print("  " + "-" * 54)
    print("  ROUTES:")
    print("  GET  /            → index.html (frontend)")
    print("  GET  /api/health  → Gemini connection check")
    print("  POST /api/step1   → FSD upload → JSON1   (1 call)")
    print("  POST /api/step2   → JSON1 → Summary+Tests (2 calls)")
    print("  POST /api/step3   → JSON1 → Gap Analysis  (1 call)")
    print("  POST /api/export  → Excel download         (0 calls)")
    print("  POST /api/chat    → AI chat per message    (1 call)")
    print("=" * 58)
    print()
    log("STARTUP", "Starting Flask server...")
    app.run(debug=False, host="0.0.0.0", port=PORT)
